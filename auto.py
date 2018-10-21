import numpy as np
import os
import openpyxl
import shutil
from cfg import *
from da_utils import *
from swmm_tools import *
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from wadi_tools import *
import datetime

filepath = "resultinho.xlsx"
if os.name == 'nt':
    slash = '\\'
else:
    slash = '//'

# converte a informacao definida no arquivo cfg.py para o formato de data/hora do python
def get_date_from_dt (dt):
    return datetime.datetime(int(dt['year']), int(dt['month']), int(dt['day']), int(dt['hour']), int(dt['minute']))

# salva os dados no excel
# data: valores a serem salvos
# swmm_date: tupla que guarda inicio da simulacao e o tamanho do timestep
# start_date_time: hora que a info foi informada
# end_date_time: instante que uma nova informação chega
def save_xlsx(data, swmm_date, start_date_time, end_date_time):
    wb = openpyxl.Workbook()
    for key, value in data.items():
        current_date_time = swmm_date[0]
        wb.create_sheet(key)
        ws = wb.get_sheet_by_name(key)
        for i, v in enumerate(value):
            ws.cell(row=i+1, column=1).value = current_date_time
            # se a data da simu é anterior a data da info, nao grave nada
            if start_date_time <= current_date_time < end_date_time:
                ws.cell(row=i+1, column=2).value = v
            else:
                ws.cell(row=i+1, column=2).value = None
            current_date_time = current_date_time + datetime.timedelta(seconds=swmm_date[1])
    wb.save(filepath)

# essa função simula os valores nos nós especificados, para o periodo especificado
# e chama a funcao para salvar no excel
def get_simulations(start_date_time, end_date_time):
    debuglog('Rodando swmm da pasta {}'.format(os.getcwd()))
    debuglog('Rodando arquivo {}'.format(swmm_filename))
    swmm = SwmmTools(swmm_filename)
    values = {}
    for node in result_nodes:
        values[node] = swmm.get_level_by_id(node)
    save_xlsx(values, (swmm.get_report_start_date(), swmm.get_report_step()), start_date_time, end_date_time)
    del swmm
# esta funcao percorre cada pasta com seus respectivos valores
# de chuva e determina o limite final daquela informação.
# O limite será até encontrar uma nova info (pasta) ou até o valor definido no swmm.
def run_and_save_folder_models():
    folders, dtlist = get_dir_names(raindata)
    owd = os.getcwd()
    for i, (folder, dt) in enumerate(zip(folders, dtlist)):
        # entra na proxima pasta
        os.chdir(folder)
        if (i != 0):
            start_date_time = get_date_from_dt(dt)
        else:
            # se for o primeiro dado informado, colocar um limite bem pequeno
            start_date_time = datetime.datetime(1900,1,1)
        if (i == len(dtlist) - 1):
            # se for o ultimo dado informado, colocar um limite de dia bem grande
            # isso faz com que o algoritmo rode ate o final da simulation
            end_date_time = datetime.datetime(3000,1,1)
        else:
            end_date_time = get_date_from_dt(dtlist[i+1])
        debuglog(str(start_date_time) + ' '+ str(end_date_time))
        get_simulations(start_date_time, end_date_time)
        # volta pra pasta raiz
        os.chdir(owd)
    return folders

def join_xlsx(folders):
    filename = "resultao.xlsx"
    wbfinal = openpyxl.Workbook()
    final_data = {}
    # percorre os nos
    for node in result_nodes:
        # cria as abas da planilha final
        wbfinal.create_sheet(node)

    # percorre as pastas
    for folder in folders:
        wbpartial = load_workbook(filename = folder+slash+filepath)
        # percorre cada no de interesse
        for node in result_nodes:
            wspartial = wbpartial.get_sheet_by_name(node)
            wsfinal = wbfinal.get_sheet_by_name(node)
            for i in range(1, wspartial.max_row+1):
                if wspartial.cell(row=i, column=2).value != None:
                    # a coluna 1 é a data/hora. A coluna 2 é o valor do nivel simulado
                    wsfinal.cell(row=i, column=1).value = wspartial.cell(row=i, column=1).value
                    wsfinal.cell(row=i, column=2).value = wspartial.cell(row=i, column=2).value
    wbfinal.save(filename)


def main():
    folders = run_and_save_folder_models()
    join_xlsx(folders)

if __name__ == "__main__":
    main()
