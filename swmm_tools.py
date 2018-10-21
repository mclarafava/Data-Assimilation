# -*- coding: utf-8 -*-
import openpyxl
import numpy as np
from swmm5.swmm5tools import SWMM5Simulation
import xlrd
from datetime import datetime, timedelta
from collections import OrderedDict
from _consts import *
from parse_swmm import *
import dateutil.parser

class SwmmTools:
    def __init__ (self, swmm_filename, data_filename = None):
        self.swmm_filename = swmm_filename
        self.data_filename = data_filename
        self.st = None

    # run a simulation if necessary
    def check_simulation(self):
        if self.st == None:
            self.run_swmm()

    # get the observed id from spreadsheet
    def get_observed_nodes (self):
        if self.data_filename == None:
            print('Function available only if a spreadsheet is defined.')
            return
        wb = openpyxl.load_workbook(self.data_filename)
        return wb.get_sheet_names()

    # get the simulated water depth by a specific id
    def get_level_by_id(self, id):
        self.check_simulation()
        return np.array(list(self.st.Results('NODE', id, depth_idx)))

    def get_report_start_date(self):
        tempnet = ParseSwmm(self.swmm_filename)
        _, dia = tempnet.get_key_by_id('OPTIONS', 'REPORT_START_DATE')
        _, hora = tempnet.get_key_by_id('OPTIONS', 'REPORT_START_TIME')
        dia = dia.split()[1]
        hora = hora.split()[1]
        diahora = dia + ' ' + hora
        return dateutil.parser.parse(diahora)

        #print(tempnet.get_contents_from_value(dia))
        #print(tempnet.get_contents_from_value(hora))


    # Careful! this function round the start_date_time to midnight
    def get_start_date(self):
        self.check_simulation()
        dia = self.st.SWMM_StartDate
        seconds = (dia - 25569) * 86400
        return datetime.utcfromtimestamp(seconds).replace(minute=0, hour =0)

    def get_report_step(self):
        self.check_simulation()
        return self.st.SWMM_ReportStep

    def get_Nperiods(self):
        self.check_simulation()
        return self.st.SWMM_Nperiods

    # recover simulated water depth from all POI
    def load_simulation_data(self):
        self.check_simulation()
        simulado_todos = np.array([],dtype=np.float)
        for id in self.get_observed_nodes():
            nivel_simulado = self.get_level_by_id(id)
            simulado_todos = np.concatenate ((simulado_todos, nivel_simulado))
        return simulado_todos

    def load_field_data(self):
        wb = openpyxl.load_workbook(self.data_filename)
        observado_todos = np.array([],dtype=np.float)
        for aba in self.get_observed_nodes():
            sheet = wb.get_sheet_by_name(aba)
            nivel_medido = np.array([], dtype=np.float)
            for i in range(2, sheet.max_row+1):
                nivel_medido = np.append(nivel_medido, sheet.cell(row=i,column=3).value)
            observado_todos = np.concatenate ((observado_todos, nivel_medido))
        observado_todos = np.array(observado_todos, dtype=np.float)
        return observado_todos

    def calc_nash(self):
        sim_data = self.load_simulation_data()
        field_data = self.load_field_data()
        num = np.nansum((field_data - sim_data)**2)
        den = np.nansum((field_data - np.nanmean(field_data))**2)
        if den == 0:
            den = 0.0000001
        nash = 1 - (num/den)
        return nash

    def run_swmm(self):
        self.st = SWMM5Simulation(self.swmm_filename) # run a simulation

def main():
    spreadsheet_name = 'data/dados.xlsx'
    swmm_filename = 'models/modificado.INP'
    tmp_dir = 'tmp'

    # cria o objeto
    calib = SwmmTools(swmm_filename, spreadsheet_name)

    # exibe os nos que possuem dados (baseado nas abas do excel)
    nodes = calib.get_observed_nodes()
    print(nodes)

    # pega os niveis simulados de um node especifico
    niveis = calib.get_level_by_id('N-29')
    print(niveis)

    # carrega os valores observados dos nodes
    field_data = calib.load_field_data()
    print(field_data)

    # carrega os valores simulados dos nodes
    sim_data = calib.load_simulation_data()
    print(sim_data)

    nash = calib.calc_nash()
    print('nash: ' + str(nash))

if __name__ == '__main__':
    main()
